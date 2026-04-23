"""
Excel I/O 工具（v1.3.0）

封装 openpyxl，提供两类操作：
- read_excel：读单元格 / 行 / 列 / 整表，输出 markdown / csv / json 文本
- append_excel：在指定 sheet 末尾追加一行或多行，缺表头时按列名顺序自动建表头

并发写保护：append 路径上加文件锁（同目录 .lock 文件 O_EXCL 互斥），
读路径不加锁（openpyxl read_only 模式可与他人写并发，最多读到旧版本）。
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import time
from typing import Optional, List, Dict, Any

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


_RANGE_CELL = re.compile(r'^([A-Za-z]+)(\d+)$')
_RANGE_AREA = re.compile(r'^([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)$')
_RANGE_COL = re.compile(r'^[A-Za-z]+$')
_RANGE_ROW = re.compile(r'^\d+$')


# ─────────────────────────────────────────────
# 读取
# ─────────────────────────────────────────────
def _normalize_path(path: str) -> str:
    """统一展开用户家目录占位符（~），相对路径保持不变（执行目录为项目根）。"""
    return os.path.expanduser(path) if path else path


def read_excel(path: str, sheet: Optional[str] = None,
               range_spec: str = "all", fmt: str = "markdown") -> str:
    """读取 xlsx 内容并返回文本表示。

    Args:
        path: xlsx 绝对路径
        sheet: sheet 名（None 表示首个 sheet）
        range_spec: 范围
            - "all" → 整张表
            - "A" / "B" / "AA" → 整列（按 Excel 列字母）
            - "1" / "2" → 整行（1-indexed）
            - "B3" → 单元格
            - "A1:C10" → 区域
        fmt: 输出格式 markdown / csv / json
    """
    _ensure_openpyxl()
    path = _normalize_path(path)
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    # data_only=True：取公式的计算结果而非公式本身
    # 不开 read_only：read_only 模式不支持 ws[spec] 区域 / 单元格直接访问
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows_2d = _slice_range(ws, range_spec)
    finally:
        wb.close()

    if not rows_2d:
        return ""
    return _format_rows(rows_2d, fmt)


def _slice_range(ws, range_spec: str) -> List[List[Any]]:
    """把 range_spec 解析为二维数据列表。"""
    spec = (range_spec or "all").strip()

    if spec.lower() == "all":
        return [list(row) for row in ws.iter_rows(values_only=True)]

    m = _RANGE_AREA.match(spec)
    if m:
        return [[c.value for c in row] for row in ws[spec.upper()]]

    m = _RANGE_CELL.match(spec)
    if m:
        return [[ws[spec.upper()].value]]

    if _RANGE_COL.match(spec):
        col_letter = spec.upper()
        col_idx = column_index_from_string(col_letter)
        out = []
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            out.append([row[0]])
        return out

    if _RANGE_ROW.match(spec):
        row_idx = int(spec)
        # iter_rows 是生成器，min/max_row 可以定位单行
        for row in ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
            return [list(row)]
        return []

    raise ValueError(f"无法解析范围表达式：{range_spec}")


def _format_rows(rows: List[List[Any]], fmt: str) -> str:
    fmt = (fmt or "markdown").lower()
    if fmt in ("md", "markdown"):
        return _to_markdown(rows)
    if fmt == "csv":
        return _to_csv(rows)
    if fmt == "json":
        return _to_json(rows)
    raise ValueError(f"不支持的格式：{fmt}（仅支持 markdown / csv / json）")


def _to_markdown(rows: List[List[Any]]) -> str:
    if not rows:
        return ""
    header = [_cell(v) for v in rows[0]]
    width = len(header)
    lines = ["| " + " | ".join(header) + " |",
             "| " + " | ".join(["---"] * width) + " |"]
    for row in rows[1:]:
        cells = [_cell(v) for v in row]
        # 行长度对齐 header
        if len(cells) < width:
            cells += [""] * (width - len(cells))
        else:
            cells = cells[:width]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def _to_csv(rows: List[List[Any]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows:
        writer.writerow([_cell(v) for v in row])
    return buf.getvalue().rstrip("\n")


def _to_json(rows: List[List[Any]]) -> str:
    if not rows:
        return "[]"
    if len(rows) == 1:
        # 单行 → JSON 数组
        return json.dumps([_cell(v) for v in rows[0]], ensure_ascii=False)
    header = [_cell(v) for v in rows[0]]
    objects = []
    for row in rows[1:]:
        obj = {}
        for i, h in enumerate(header):
            obj[h] = _cell(row[i]) if i < len(row) else ""
        objects.append(obj)
    return json.dumps(objects, ensure_ascii=False, indent=2)


def _cell(value: Any) -> str:
    """统一单元格转字符串（None → ''；管道符转义防破坏 markdown 表格）。"""
    if value is None:
        return ""
    s = str(value)
    return s.replace("|", "\\|").replace("\n", " ")


# ─────────────────────────────────────────────
# 追加
# ─────────────────────────────────────────────
def append_excel(path: str, sheet: str, rows: List[Dict[str, str]],
                 auto_create_header: bool = True) -> int:
    """向 xlsx 指定 sheet 追加行（带文件锁）。

    每行是 dict: {列名: 值}。如果 sheet 不存在或表头不存在，根据 auto_create_header
    决定是否按本次首行的列名顺序创建表头并新建 sheet。

    返回实际写入的行数。
    """
    _ensure_openpyxl()
    if not rows:
        return 0

    path = _normalize_path(path)
    with _file_lock(path):
        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            if not auto_create_header:
                raise FileNotFoundError(f"文件不存在且未启用自动建表头：{path}")
            wb = Workbook()
            # 删掉默认 sheet（之后按 sheet 名新建）
            default = wb.active
            wb.remove(default)

        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            if not auto_create_header:
                raise KeyError(f"sheet 不存在且未启用自动建表头：{sheet}")
            ws = wb.create_sheet(sheet)

        # 表头处理：以首行作为列名权威
        existing_header = _read_header(ws)
        if not existing_header:
            if not auto_create_header:
                raise ValueError(f"sheet [{sheet}] 无表头且未启用自动建表头")
            # 用本次写入第一行的列顺序当表头
            header = list(rows[0].keys())
            # 显式写到第 1 行，避免 openpyxl 在新建 sheet 时残留的占位空行
            for col_idx, name in enumerate(header, start=1):
                ws.cell(row=1, column=col_idx, value=name)
            existing_header = header
            next_row = 2
        else:
            # 已有表头：用 ws.max_row + 1 作为下一行（max_row 在有数据时是准确的）
            next_row = ws.max_row + 1
            # 防御：若 max_row 仍为 1 且该行就是表头，next_row=2 也是对的

        col_index = {name: i for i, name in enumerate(existing_header)}
        written = 0
        for row in rows:
            for k, v in row.items():
                if k in col_index:
                    ws.cell(row=next_row, column=col_index[k] + 1,
                            value=v if v is not None else "")
            next_row += 1
            written += 1

        wb.save(path)
        return written


def _read_header(ws) -> List[str]:
    """读取首行作为表头（空行返回空列表）"""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        # 全空算无表头
        if not any(cells):
            return []
        # 末尾连续空裁掉
        while cells and cells[-1] == "":
            cells.pop()
        return cells
    return []


# ─────────────────────────────────────────────
# 文件锁
# ─────────────────────────────────────────────
class _file_lock:
    """同目录 .lock 文件 O_EXCL 互斥锁，重试 5 秒。"""
    def __init__(self, target_path: str, retries: int = 25, retry_interval: float = 0.2):
        self.lock_path = target_path + ".lock"
        self.retries = retries
        self.retry_interval = retry_interval
        self._fd = None

    def __enter__(self):
        # 确保父目录存在（首次 append 全新文件时 target 还不存在）
        parent = os.path.dirname(self.lock_path)
        if parent and not os.path.exists(parent):
            os.makedirs(parent, exist_ok=True)
        for _ in range(self.retries):
            try:
                self._fd = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                return self
            except FileExistsError:
                time.sleep(self.retry_interval)
        raise TimeoutError(f"等待文件锁超时：{self.lock_path}")

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            if self._fd is not None:
                os.close(self._fd)
        except Exception:
            pass
        try:
            if os.path.exists(self.lock_path):
                os.remove(self.lock_path)
        except Exception:
            pass
        return False


def _ensure_openpyxl():
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl 未安装。请执行 `pip install openpyxl>=3.1` 后再使用 Excel 步骤。"
        )
